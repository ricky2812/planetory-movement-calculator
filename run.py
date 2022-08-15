#python run.py file_name.txt
from openpyxl import load_workbook, Workbook
from openpyxl.styles.colors import Color
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import sys



nakshatra_config = {
    "ashwini": "ashwini", "bharani": "bharni", "krithika": "krithika", "rohini": "rohini", "mrigasira": "mrigashira",
    "ardra": "aaradha", "punarvasu": "punarvasu", "pushyami": "pushya", "ashlesha": "aslesha", "magha": "magha",
    "poorva phalguni": "poorva_phalguni", "uttara phalguni": "uttara_phalguni", "hasta": "hasta", "chitra": "chitra",
    "swathi": "swati", "vishakha": "vishaka", "anuradha": "anooradha", "jyeshta": "jyeshtha", "moola": "moola",
    "poorva ashadha": "poorva_ashadha", "uttara ashadha": "uttara_ashadha", "shravana": "sravana", "dhanishta": "dhanishta",
    "shatabhisha": "satabhisak", "poorva bhadrapada": "poorva_bhadra", "uttra bhadrapada": "uttara_bhadra", "revathi": "revati"}

rasi_dict = {'Ar': 'Aries', 'Ta': 'Taurus', 'Ge': 'Gemini', 'Cn': 'Cancer', 'Le': 'Leo', 'Vi': 'Virgo',
             'Li': 'Libra', 'Sc': 'Scorpio', 'Sg': 'Sagittarius', 'Cp': 'Capricorn', 'Aq': 'Aquarius', 'Pi': 'Pisces'}

nakshatra_dict = {'Aswi': 'Ashwini', 'Bhar': 'Bharani', 'Krit': 'Krithika', 'Rohi': 'Rohini', 'Mrig': 'Mrigasira',
                  'Ardr': 'ardra', 'Puna': 'Punarvasu', 'Push': 'Pushyami', 'Asre': 'Ashlesha', 'Magh': 'Magha',
                  'PPha': 'Poorva Phalguni', 'UPha': 'Uttara Phalguni', 'Hast': 'Hasta', 'Chit': 'Chitra',
                  'Swat': 'Swathi', 'Visa': 'Vishakha', 'Anu': 'Anuradha',	'Jye': 'Jyeshta', 'Mool': 'Moola',
                  'PSha': 'Poorva Ashadha', 'USha': 'Uttara Ashadha', 'Srav': 'Shravana', 'Dhan': 'Dhanishta',
                  'Sata': 'Shatabhisha', 'PBha': 'poorva Bhadrapada', 'UBha': 'Uttra Bhadrapada', 'Reva': 'Revathi'}

months = ['January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December']

ordered_list = ["Date", "Time", "Planets", "Nakshatra", "Pada","Rasi","Longitudes", "Lagna", "Sun", "Moon", "Mars", "Mercury",
                        "Jupiter", "Venus", "Saturn", "Rahu", "Ketu"]


def calculate_house(planet_nakshatra, pada):
    """
    i/p parameters
    planet_nakshatra : String
    pada : String
    o/p
    List of plotted houses
    """
    houses = list()
    house_wise = []
    # Generate sub houses for each nakshatra
    # each house has 4 values
    for val in nakshatra_config:
        for j in range(1, 5):
            houses.append(nakshatra_config[val]+"_"+str(j))
    # get the position of the planet in the house
    sub_house = nakshatra_config[planet_nakshatra.lower().strip()] + "_" + pada
    ind = houses.index(sub_house)
    if ind is None:
        return "Wrong Sub House"
    # based on the plotting start from the middle of the list and rotate the whole array in circular manner
    i = ind
    n = len(houses)
    # logic for the circular array
    while i < n + ind:
        house_wise.append(houses[(i % n)])
        i = i + 1
    # divide the houses in terms of 9 paths
    n = 9
    final = [house_wise[i * n:(i + 1) * n]
             for i in range((len(house_wise) + n - 1) // n)]
    return final


def get_sub_house_data(house_data, sub_house_data):
    """
    i/p parameters
    house_data: List
    sub_house_data: String
    """
    calc_houses = dict()
    # this will plot the houses
    for i in range(1, 13):
        calc_houses[str(i)] = house_data[i-1]
    # get the position of the sub house
    for data in calc_houses:
        for val in calc_houses[data]:
            if val == sub_house_data:
                return data+"."+str(calc_houses[data].index(val))


def write_to_excel(final_list):
    """
    i/p parameters
    final_dict: Dictionary
    """
    filename = "output.xlsx"
    try:
        # try to open existing file and append to it
        wb = load_workbook(filename)
        for sheet in final_list:
            for key, value in sheet.items():
                ws = wb[key]
                ws.append(value)
    except:
        # create a new file if it does not exist
        wb = Workbook()
        for sheet in final_list:
            for key, value in sheet.items():
                ws = wb.create_sheet(key)
                ws.append(ordered_list)
                ws.append(value)
        wb.remove(wb['Sheet'])

    #This is the part where the style formatting gets done
    color1 = PatternFill(start_color='DBF3FA', end_color='DBF3FA', fill_type='solid')
    color2 = PatternFill(start_color='7AD7F0', end_color='7AD7F0', fill_type='solid')
    for sheet in wb:
        ws = wb[sheet.title]
        for row in ws.iter_rows():
            for cell in row:
                cell.fill = color1
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(border_style='thin', color='000000'),
                                     right=Side(border_style='thin', color='000000'),
                                     top=Side(border_style='thin', color='000000'),
                                     bottom=Side(border_style='thin', color='000000'))
                c = cell.coordinate  
                if c[0] in ['A', 'B', 'C', 'E', 'F']:
                    cell.font=Font(bold=True)
                    cell.fill = color2
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif c[0] in ['D', 'G']:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.font=Font(bold=True)
                    cell.fill = color2
                    
                try:
                    if cell.value == '0':
                        cell.font=Font(bold=True, color=Color('FF0000'))
                        cell.fill = color1
                    elif '.' in cell.value:
                            final_str = cell.value.split('.')
                            if final_str[1] == '0':
                                    cell.value = int(final_str[0])
                                    cell.font=Font(bold=True, color=Color('FF0000'))
                                    cell.fill = color1
                            elif final_str[1] in ['1', '2', '3', '4', '5', '6', '7', '8', '9']:
                                    cell.value = float(cell.value)
                except TypeError:
                    pass
    wb.save(filename)

def get_datetime(data):
    date=dict()
    fulldate, time=data[2:4]
    fulldate=fulldate.split()[1:]
    time=time.split()[1].split(':')
    date['year']=fulldate[2]
    date['month']=months.index(fulldate[0]) + 1
    date['date']=fulldate[1].strip(',')
    date['hour']=time[0]
    date['minute']=time[1]
    date['second']=time[2]
    return date

def get_info(item):
    data=item.split()
    body=data[0]
    data=data[-8:]
    planet=dict()
    planet['planet']=body
    planet['longitude']=data[0] + ' ' + data[1] + ' ' + data[2] + ' ' + data[3]
    planet['nakshatra']=nakshatra_dict[data[-4]]
    planet['pada']=data[-3]
    planet['rasi']=rasi_dict[data[-2]]
    return planet

def main():
    final_list=list()
    with open(sys.argv[1], 'r') as f:
        data=f.readlines()
    for line in range(len(data)):
        data[line]=data[line].strip()

    data=[x for x in data if x]
    data_date=get_datetime(data)
    info=data[45:55]
    planetary_data=dict()
    for item in info:
        planetary_data[item.split()[0]]=get_info(item)
    for name in ["Lagna", "Sun", "Moon", "Mars", "Mercury", "Jupiter", "Venus", "Saturn", "Rahu", "Ketu"]:
        # get the data for the planet
        date=data_date['date'] + '/' + \
            str(data_date['month']) + '/' + data_date['year']
        time=data_date['hour'] + ":" + \
            data_date['minute'] + ":" + data_date['second']
        data_for_excel=dict()
        # for name in planets:
        nakshatra=planetary_data[name]['nakshatra']
        pada=str(planetary_data[name]['pada'])
        house_data=calculate_house(nakshatra, pada)
        if name not in data_for_excel:
            data_for_excel[name]=list()
            data_for_excel[name].append(date)
            data_for_excel[name].append(time)
            data_for_excel[name].append(name)
            data_for_excel[name].append(nakshatra)
            data_for_excel[name].append(pada)
            data_for_excel[name].append(planetary_data[name]['rasi'])
            data_for_excel[name].append(planetary_data[name]['longitude'])
        for data in planetary_data:
            nakshatra=planetary_data[data]['nakshatra']
            pada=str(planetary_data[data]['pada'])
            sub_house_data=nakshatra_config[nakshatra.lower(
            ).strip()] + "_"+pada
            data_temp = get_sub_house_data(house_data, sub_house_data)
            data_for_excel[name].append(data_temp)
        final_list.append(data_for_excel)
    write_to_excel(final_list)

if __name__ == "__main__":
    main()
    print("Success!!!")

